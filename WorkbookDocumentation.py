""" processing of Tableau workbooks to extract metadata for documentation """
import os
import time
import logging
import json
import openpyxl
from openpyxl.styles import Font
import Handle_twbx
from validator.validate_styles import validate_styles

# from tkinter import messagebox


class WorkbookDocumentation:
    """Core workbook class with methods to extract metadata"""

    def __init__(self, input_file, style_guide=None):
        if isinstance(input_file, str):
            workbook_tree, workbook_contents = Handle_twbx.xml_open(input_file)
        else:
            workbook_tree = input_file
        self.style_guide = style_guide
        self.root = workbook_tree
        self.input_file = input_file
        self.connections = []
        self.tables = []
        self.parameters = []
        self.custom_sql_queries = []
        self.calculations = []
        self.columns = []
        self.sets = []
        self.styles = []

        self.document_type = workbook_tree.getroot().tag

        if self.document_type == "datasource":
            self.datasource_root = workbook_tree.getroot()
            self.process_datasource(workbook_tree.getroot())
        else:
            self.datasource_root = workbook_tree.find(".//datasources")
            try:
                for datasource_node in self.datasource_root.findall("./datasource"):
                    self.process_datasource(datasource_node)
            except AttributeError:
                print("No data sources found")

            self.worksheet_root = workbook_tree.find(".//worksheets")
            self.worksheet_columns = []
            self.worksheet_captions = []
            try:
                for worksheet_node in self.worksheet_root.findall("./worksheet"):
                    self.find_worksheet_columns(worksheet_node)
                    self.find_worksheet_captions(worksheet_node)
            except AttributeError:
                print("No worksheets found")

            self.dashboard_root = workbook_tree.find(".//dashboards")
            self.dashboard_objects = []
            try:
                for dashboard_node in self.dashboard_root.findall("./dashboard"):
                    self.find_dashboards(dashboard_node)
            except AttributeError:
                print("No dashboards found")

        if style_guide is not None:
            style_guide_json = self.ingest_style_guide()
            style_guide_json.pop("_README")
            # workbook_file = self.ingest_tableau_workbook()
            # self.styles = validate_styles(style_guide_json, workbook_file)
            self.styles = validate_styles(style_guide_json, workbook_contents)

        self.out_file = ""

    def ingest_tableau_workbook(self):
        """Ingest Tableau Workbook file (~/foo.twb) from command line arguments."""

        # Pass Tableau Workbook to parser as open file
        with open(
            self.input_file, mode="r", encoding="utf-8"
        ) as tableau_workbook_infile:
            tableau_workbook_file = tableau_workbook_infile.read()

        return tableau_workbook_file

    def ingest_style_guide(self):
        """Ingest JSON style guide file (~/foo.json) from command line arguments."""

        # Test Style Guide input for valid JSON
        with open(self.style_guide, mode="r", encoding="utf-8") as style_guide_infile:
            try:
                style_guide_json = json.load(style_guide_infile)

            except json.JSONDecodeError:
                print(
                    "Invalid JSON format. \n"
                    "Check for double quotes and matching brackets."
                )

        return style_guide_json

    def process_datasource(self, datasource_node):
        """iterate through each data source and find information"""
        if "caption" in datasource_node.attrib:
            datasource_name = datasource_node.attrib["caption"]
        elif "formatted-name" in datasource_node.attrib:
            datasource_name = datasource_node.attrib["formatted-name"]
        else:
            # todo: should there be a fallback if none of the 3 attributes exist?
            datasource_name = datasource_node.attrib["name"]
        logging.info("now processing %s data source", datasource_name)
        self.find_connections(datasource_node, datasource_name)
        self.find_parameters(datasource_node, datasource_name)
        self.find_calculations(datasource_node, datasource_name)
        self.find_custom_sql(datasource_node, datasource_name)
        self.find_tables(datasource_node, datasource_name)
        self.find_sets(datasource_node, datasource_name)
        if datasource_node.findall(
            "./connection/named-connections/named-connection/"
            + "connection[@class='excel-direct']"
        ) or datasource_node.findall("./connection[@class='excel-direct']"):
            logging.info("found Excel. Skipping columns")
            # todo Extract columns from single connection Excel data source
        else:
            self.find_columns(datasource_node, datasource_name)

    @staticmethod
    def _validate_attribute_(node, attribute):
        """validate that the attribute exists in the node"""
        if node is not None:
            if attribute in node.attrib:
                value = node.attrib[attribute]
            else:
                value = ""
        else:
            value = ""
        return value

    # @staticmethod
    def _resolve_calculations(self, calculation_list: list) -> list:
        """iterate through calculations and replace names that are found with the underlying calculation"""
        for calculation in calculation_list:
            calc_resolved = calculation["calculation"]
            # if calc_resolved.find("[Calculation_55450621892177920]") != -1:
            #     print("Found it")
            self._replace_calc_names_with_formulas(calc_resolved, calculation)

        return calculation_list

    def _replace_calc_names_with_formulas(self, calc_resolved, calculation, depth=0):
        """Recursive replace Tableau internal names with formulas"""
        changed = False
        for replacement in self.calculations:
            # find if the replacement value is in the string and be sure that the replacement doesn't equal what
            # it's replacing. If the values match, there will be infinite recursion. That's bad.
            if calc_resolved.find(replacement["name"]) != -1 and not (
                (
                    replacement["datasource"] == "Parameters"
                    and replacement["name"] == f"[{replacement['caption']}]"
                )
                or replacement["name"] == replacement["calculation"]
            ):
                changed = True
                if replacement["datasource"] == "Parameters":
                    calc_resolved = calc_resolved.replace(
                        replacement["name"], f"[{replacement['caption']}]"
                    )
                else:
                    calc_resolved = calc_resolved.replace(
                        replacement["name"], replacement["calculation"]
                    )
        calculation["calc_resolved"] = calc_resolved

        # recurse if any changes made to make sure none have been missed
        if changed:
            depth += 1
            self._replace_calc_names_with_formulas(calc_resolved, calculation, depth)
        return calculation

    # @staticmethod
    def _resolve_names(self, calculation_list: list) -> list:
        """iterate through calculations name column and replace names that are found with the caption (UI name)"""
        for calculation in calculation_list:
            name_resolved = calculation["name"]
            for replacement in self.calculations:
                name_resolved = name_resolved.replace(
                    replacement["name"], replacement["caption"]
                )
            calculation["name_resolved"] = name_resolved
        return calculation_list

    # @staticmethod
    def _resolve_names_in_calcs(self, calculation_list: list) -> list:
        """iterate through calculations and replace names that are found with the caption"""
        for calculation in calculation_list:
            name_resolved = calculation["calculation"]
            for replacement in self.calculations:
                name_resolved = name_resolved.replace(
                    replacement["name"], f"[{replacement['caption']}]"
                )
            calculation["calc_renamed"] = name_resolved
        return calculation_list

    def find_connections(self, datasource_node, datasource_name):
        """iterate through connection nodes to find data"""
        spreadsheet_columns = ["datasource", "connection", "type"]
        for connection_node in datasource_node.findall("./connection"):
            # find all federated connections and their data
            if connection_node.find("./named-connections"):
                for named_connection_node in connection_node.findall(
                    "./named-connections/named-connection"
                ):
                    if (
                        self._validate_attribute_(named_connection_node, "caption")
                        != ""
                    ):
                        connection_name = self._validate_attribute_(
                            named_connection_node, "caption"
                        )
                    else:
                        connection_name = self._validate_attribute_(
                            named_connection_node, "name"
                        )

                    fed_connection_xpath = "./connection"
                    for federated_connection_node in named_connection_node.findall(
                        fed_connection_xpath
                    ):
                        connection_type = self._validate_attribute_(
                            federated_connection_node, "class"
                        )
                        # print("Connection type: " + connection_type)
                        connection_values = [
                            datasource_name,
                            connection_name,
                            connection_type,
                        ]
                        self.connections.append(
                            dict(zip(spreadsheet_columns, connection_values))
                        )
                        # print(self.connections)
            # find all server data sources (and other non-federated types) and their data
            else:
                connection_type = self._validate_attribute_(connection_node, "class")
                connection_values = [
                    datasource_name,
                    self._validate_attribute_(connection_node, "dbname"),
                    connection_type,
                ]
                self.connections.append(
                    dict(zip(spreadsheet_columns, connection_values))
                )
        logging.info("Found %s connections", str(len(self.connections)))

    def find_tables(self, datasource_node, datasource_name):
        """iterate through relation nodes to find tables"""
        # todo More efficient to do this within each connection?
        spreadsheet_columns = ["datasource", "connection", "name", "table"]
        for node in datasource_node.findall(".//relation[@type='table']"):
            if node.attrib["name"] != "Extract" and node.attrib["name"] != "sqlproxy":
                # if 'connection' in node.attrib:
                table_values = [
                    datasource_name,
                    self._validate_attribute_(node, "connection"),
                    self._validate_attribute_(node, "name"),
                    self._validate_attribute_(node, "table"),
                ]
                self.tables.append(dict(zip(spreadsheet_columns, table_values)))
        logging.info("Found %s tables", str(len(self.tables)))

    def find_custom_sql(self, datasource_node, datasource_name):
        """iterate through relation nodes to find custom SQL"""
        # todo More efficient to do this within each connection?
        spreadsheet_columns = ["datasource", "connection", "name", "SQL"]
        for node in datasource_node.findall(".//relation[@type='text']"):
            if node.attrib["name"] != "Extract":
                # print(node.attrib["connection"] + ',' + node.attrib["name"] + ',' + node.text)
                custom_sql_values = [
                    datasource_name,
                    self._validate_attribute_(node, "connection")
                    + ","
                    + self._validate_attribute_(node, "name")
                    + ","
                    + node.text,
                ]
                self.custom_sql_queries.append(
                    dict(zip(spreadsheet_columns, custom_sql_values))
                )
        logging.info("Found %s custom SQL queries", str(len(self.custom_sql_queries)))

    def find_parameters(self, datasource_node, datasource_name):
        """iterate through data sources nodes to find Parameter source"""
        spreadsheet_columns = [
            "datasource",
            "caption",
            "value",
            "datatype",
            "type",
            "role",
            "name",
            "description",
        ]
        for node in datasource_node.findall(".[@hasconnection='false']/column"):
            parameter_values = [
                datasource_name,
                self._validate_attribute_(node, "caption"),
                self._validate_attribute_(node, "value"),
                self._validate_attribute_(node, "datatype"),
                self._validate_attribute_(node, "type"),
                self._validate_attribute_(node, "role"),
                self._validate_attribute_(node, "name"),
            ]
            description = self._get_description(node)
            parameter_values.append(description)
            self.parameters.append(dict(zip(spreadsheet_columns, parameter_values)))
        logging.info("Found %s parameters", str(len(self.parameters)))

    @staticmethod
    def _get_description(node):
        description = ""
        for desc_node in node.findall(".//run"):
            description = description + " " + desc_node.text
        description = description.lstrip()
        return description

    def find_columns(self, datasource_node, datasource_name):
        """iterate through column nodes to find columns"""
        # todo More efficient to do this within each connection?
        spreadsheet_columns = [
            "datasource",
            "key",
            "table",
            "column",
            "caption",
            "datatype",
            "hidden",
            "description",
        ]
        for node in datasource_node.findall("./column"):
            if node.find("./calculation") is None and node.find("./aliases") is None:
                # print(node.attrib["name"] + " is not a calculation")
                name = node.attrib["name"]
                caption = self._validate_attribute_(node, "caption")
                datatype = self._validate_attribute_(node, "datatype")
                hidden = self._validate_attribute_(node, "hidden")
                column_node = self.datasource_root.find(
                    ".//connection/cols/map[@key='" + name + "']"
                )
                # print("./connection/cols/map[@key='" + name + "']")
                if column_node is not None:
                    table, column = column_node.attrib["value"].split("].[")
                    table = table + "]"
                    column = "[" + column
                else:
                    table = ""
                    column = ""

                description = self._get_description(node)

                column_values = [
                    datasource_name,
                    name,
                    table,
                    column,
                    caption,
                    datatype,
                    hidden,
                    description,
                ]
                self.columns.append(dict(zip(spreadsheet_columns, column_values)))
        logging.info("Found %s columns", str(len(self.columns)))

    def find_calculations(self, datasource_node, datasource_name):
        """iterate through column nodes to find calculations"""
        # todo More efficient to do this within each connection?
        spreadsheet_columns = [
            "datasource",
            "caption",
            "name",
            "role",
            "calculation_type",
            "hidden",
            "datatype",
            "default_format",
            "calculation",
            "description",
        ]
        for node in datasource_node.findall("./column[@caption][calculation]"):
            # print(node.attrib)
            caption = node.attrib["caption"]
            name = node.attrib["name"]
            role = node.attrib["role"]
            calculation_type = node.attrib["type"]

            datatype = self._validate_attribute_(node, "datatype")

            default_format = self._validate_attribute_(node, "default-format")
            hidden = self._validate_attribute_(node, "hidden")
            calc_node = node.find("./calculation")
            calculation = self._validate_attribute_(calc_node, "formula")

            description = self._get_description(node)
            calculation_values = [
                datasource_name,
                caption,
                name,
                role,
                calculation_type,
                hidden,
                datatype,
                default_format,
                calculation,
                description,
            ]
            self.calculations.append(dict(zip(spreadsheet_columns, calculation_values)))

        # Resolve reference to other calculations. Parameters are excluded since they are variable
        self.calculations = self._resolve_calculations(self.calculations)
        self.calculations = self._resolve_names_in_calcs(self.calculations)

        logging.info("Found %s calculations", str(len(self.calculations)))

    def find_sets(self, datasource_node, datasource_name):
        """iterate through datasource node to find sets"""
        spreadsheet_columns = [
            "datasource",
            "caption",
            "name",
            "element",
            "type",
            "condition_calculation",
            "number",
            "end",
            "direction",
            "members",
            "expression",
            "description",
        ]

        for set_node in datasource_node.findall(
            "./group[@{http://www.tableausoftware.com/xml/user}ui-builder='filter-group']"
        ):
            caption = self._validate_attribute_(set_node, "caption")
            name = self._validate_attribute_(set_node, "name")
            element = ""
            set_type = ""
            condition_calculation = ""
            number = ""
            end = ""
            direction = ""
            expression = ""
            members = []
            for groupfilter_node in set_node.findall("./groupfilter"):
                function = groupfilter_node.attrib["function"]
                if function == "union":
                    for node in groupfilter_node.findall("./groupfilter"):
                        if node.attrib["function"] == "member":
                            members.append(node.attrib["member"])
                            element = node.attrib["level"]
                            set_type = "manual selection"
                        elif node.attrib["function"] == "reference":
                            members.append(node.attrib["field"])
                            set_type = "combined set"

                elif function == "filter":
                    condition_calculation = groupfilter_node.attrib["expression"]
                    for node in groupfilter_node.findall("./groupfilter"):
                        element = node.attrib["level"]
                    set_type = "condition"
                elif function == "end":
                    end = groupfilter_node.attrib["end"]
                    number = groupfilter_node.attrib["count"]
                    for calc_node in groupfilter_node.findall("./groupfilter"):
                        direction = calc_node.attrib["direction"]
                        condition_calculation = calc_node.attrib["expression"]
                        for element_node in calc_node.findall("./groupfilter"):
                            function = element_node.attrib["function"]
                            if function == "level-members":
                                element = element_node.attrib["level"]
                            else:
                                expression = element_node.attrib["expression"]
                                for base_node in element_node.findall("./groupfilter"):
                                    element = base_node.attrib["level"]
                    set_type = "top N"

            member_list = "|".join(members)
            description = self._get_description(set_node)
            set_values = [
                datasource_name,
                caption,
                name,
                element,
                set_type,
                condition_calculation,
                number,
                end,
                direction,
                member_list,
                expression,
                description,
            ]
            self.sets.append(dict(zip(spreadsheet_columns, set_values)))

    def find_worksheet_columns(self, worksheet_node):
        """iterate through worksheets to find columns reference on them"""

        worksheet_name = worksheet_node.attrib["name"]
        start_length = len(self.worksheet_columns)
        spreadsheet_columns = [
            "worksheet",
            "datasource",
            "caption",
            "name",
            "role",
            "datatype",
            "type",
            "calculation",
            "computation",
        ]
        for datasource_dependency_node in worksheet_node.findall(
            ".//datasource-dependencies"
        ):
            # print(node.attrib)
            datasource = datasource_dependency_node.attrib["datasource"]
            for datasource_node in worksheet_node.findall(".//datasources/datasource"):
                if (
                    datasource == datasource_node.attrib["name"]
                    and "caption" in datasource_node.attrib
                ):
                    datasource = datasource_node.attrib["caption"]
            for column_node in datasource_dependency_node.findall("./column"):
                name = column_node.attrib["name"]
                # print(worksheet_name + ", " + datasource + ", " + name)
                caption = self._validate_attribute_(column_node, "caption")
                role = self._validate_attribute_(column_node, "role")
                column_type = self._validate_attribute_(column_node, "type")
                if column_type == "ordinal":
                    column_type = "continuous"
                elif column_type == "nominal":
                    column_type = "discrete"

                datatype = self._validate_attribute_(column_node, "datatype")

                calc_node = column_node.find(".//calculation")
                calculation = self._validate_attribute_(calc_node, "formula")

                if name.find("'") == -1:
                    searchstring = f"./column-instance[@column='{name}']"
                    # print(searchstring)
                else:
                    searchstring = f'./column-instance[@column="{name}"]'
                    # print(searchstring)

                computation_node = datasource_dependency_node.findall(searchstring)
                computation = None
                for column_instance in computation_node:
                    computation = self._validate_attribute_(
                        column_instance, "derivation"
                    )
                    if self._validate_attribute_(column_instance, "type"):
                        raw_type = self._validate_attribute_(column_instance, "type")
                        if raw_type == "quantitative":
                            column_type = "continuous"
                        elif raw_type == "nominal":
                            column_type = "discrete"
                    if computation != "None":
                        role = "measure"

                worksheet_values = [
                    worksheet_name,
                    datasource,
                    caption,
                    name,
                    role,
                    datatype,
                    column_type,
                    calculation,
                    computation,
                ]
                self.worksheet_columns.append(
                    dict(zip(spreadsheet_columns, worksheet_values))
                )

        self.worksheet_columns = self._resolve_calculations(self.worksheet_columns)
        self.worksheet_columns = self._resolve_names(self.worksheet_columns)
        logging.info(
            "Found %s columns in %s",
            str(len(self.worksheet_columns) - start_length),
            worksheet_name,
        )

    def find_worksheet_captions(self, worksheet_node):
        """iterate through worksheets to find captions"""

        worksheet_name = worksheet_node.attrib["name"]
        start_length = len(self.worksheet_columns)
        spreadsheet_columns = ["worksheet", "caption"]
        caption = ""
        for caption_node in worksheet_node.findall(
            "./layout-options/caption/formatted-text/run"
        ):
            caption = caption + " " + caption_node.text.replace("Æ", "")

        caption = caption.lstrip()
        if caption != "":
            worksheet_captions = [worksheet_name, caption]
            self.worksheet_captions.append(
                dict(zip(spreadsheet_columns, worksheet_captions))
            )
        logging.info(
            "Found %s captions in %s",
            str(len(self.worksheet_captions) - start_length),
            worksheet_name,
        )

    def find_dashboards(self, dashboard_node):
        """iterate through dashboard nodes to find worksheets and filters"""
        dashboard_name = dashboard_node.attrib["name"]
        spreadsheet_columns = ["dashboard", "dashboard_object", "type"]
        for node in dashboard_node.findall(".//zone[@name]"):
            # print(node.attrib)

            if "type" in node.attrib:
                # worksheets have no 'type' attribute. Text boxes, filters, and legends have 'type' attributes.
                object_type = "filter"
                # print(".//style-rule[@element='quick-filter']/format[@field='" + node.attrib["param"] + "']")

                # Added a check to see if the objects have an embedded single quote
                # The current code will break if the object name has both single and double quotes
                if node.attrib["param"].find("'") == -1:
                    searchstring = (
                        ".//style-rule[@element='quick-filter']/format[@field='"
                        + node.attrib["param"]
                        + "']"
                    )
                    # print(searchstring)
                else:
                    searchstring = (
                        './/style-rule[@element="quick-filter"]/format[@field="'
                        + node.attrib["param"]
                        + '"]'
                    )
                    # print(searchstring)

                # find the quickfilter node that has the name of the object
                name_node = self.root.find(searchstring)

                # found
                if name_node is not None:
                    # print(name_node.attrib)
                    if "value" in name_node.attrib:
                        # print("value = {}".format(name_node.attrib["value"]))
                        dashboard_object = name_node.attrib["value"]
                    else:
                        # print("no value param = {}".format(node.attrib["name"]))
                        dashboard_object = node.attrib["name"]

                # object must not be a quickfilter, look to find the value in the datasources on the worksheet
                # todo: decide how to handle these objects and document them. Currently pulls info, but doesn't save
                else:
                    try:
                        # this section is all experimental.
                        # The try/except is just prevent errors from stopping the whole thing
                        # todo: remove try/except or at least have real error handling
                        print(f"none param = {format(node.attrib['name'])}")
                        # name_node = self.root.find(".//worksheet[@name='" + node.attrib["name"] +
                        #                            "']/table/view/datasource-dependencies[column-instance/@name=")
                        print(
                            ".//worksheet[@name='"
                            + node.attrib["name"]
                            + "']/table/view/datasource-dependencies[column-instance][@name='"
                            + node.attrib["param"].split("].")[1]
                            + "']"
                        )

                        # Measure names objects have another layer of abstraction
                        # before the database dependency can be found
                        if node.attrib["param"].split("].")[1] == "[:Measure Names]":
                            filter_name = "Measure Names"

                        else:
                            print(f'Not Color: {node.attrib["param"].split("].")[1]}')
                            datasource_dependency_node = self.root.find(
                                ".//worksheet[@name='"
                                + node.attrib["name"]
                                + "']/table/view/datasource-dependencies/column-instance[@name='"
                                + node.attrib["param"].split("].")[1]
                                + "']/.."
                            )
                            print(datasource_dependency_node.attrib["datasource"])
                            column_name = datasource_dependency_node.find(
                                "./column-instance[@name='"
                                + node.attrib["param"].split("].")[1]
                                + "'][@column]"
                            ).attrib["column"]
                            print(column_name)
                            filter_node = datasource_dependency_node.find(
                                "./column[@name='" + column_name + "']"
                            )
                            if "caption" in filter_node.attrib:
                                filter_name = filter_node.attrib["caption"]
                            else:
                                filter_name = column_name

                        print(filter_name)
                    except:
                        print("Error in filter stuff")

                    dashboard_object = node.attrib["name"]
            else:
                object_type = "worksheet"
                # print("no type name = {}".format(node.attrib["name"]))
                dashboard_object = node.attrib["name"]

            dashboard_values = [dashboard_name, dashboard_object, object_type]
            self.dashboard_objects.append(
                dict(zip(spreadsheet_columns, dashboard_values))
            )

        logging.info("Found %s dashboards", str(len(self.dashboard_objects)))

    def write_documentation(self, output_dir):
        """output individual object type information to separate sheets in an Excel workbook"""
        logging.info("Writing to %s", self.out_file)
        wb = self._init_workbook()
        self._write_openpyxl_worksheet(wb, self.connections, "Connections")
        self._write_openpyxl_worksheet(wb, self.parameters, "Parameters")
        self._write_openpyxl_worksheet(wb, self.tables, "Tables")
        self._write_openpyxl_worksheet(wb, self.custom_sql_queries, "Custom SQL")
        self._write_openpyxl_worksheet(wb, self.columns, "Columns")
        self._write_openpyxl_worksheet(wb, self.calculations, "Calculations")
        self._write_openpyxl_worksheet(wb, self.sets, "Sets")
        self._write_openpyxl_worksheet(wb, self.styles, "Style validation")
        if self.document_type == "workbook":
            self._write_openpyxl_worksheet(
                wb, self.worksheet_captions, "Worksheet Captions"
            )
            self._write_openpyxl_worksheet(
                wb, self.worksheet_columns, "Worksheet Columns"
            )
            # print(self.worksheet_columns)
            worksheet_dimensions = list(
                filter(lambda x: x["role"] == "dimension", self.worksheet_columns)
            )
            self._write_openpyxl_worksheet(
                wb, worksheet_dimensions, "Worksheet Dimensions"
            )
            worksheet_dimensions = list(
                filter(lambda x: x["role"] == "measure", self.worksheet_columns)
            )
            self._write_openpyxl_worksheet(
                wb, worksheet_dimensions, "Worksheet Measures"
            )
            self._write_openpyxl_worksheet(
                wb, self.dashboard_objects, "Dashboard Objects"
            )
        self._save_workbook(wb, self.input_file, output_dir)

    @staticmethod
    def _init_workbook():
        wb = openpyxl.Workbook()
        wb.remove(wb["Sheet"])
        return wb

    @staticmethod
    def _save_workbook(wb, input_file, output_dir):
        out_file = (
            output_dir
            + os.sep
            + os.path.splitext(os.path.basename(input_file))[0]
            + " Documentation alt.xlsx"
        )
        wb.save(out_file)

    @staticmethod
    def _write_openpyxl_worksheet(wb, extracted_data, worksheet_name):
        curr_sheet = wb.create_sheet(worksheet_name)
        if len(extracted_data) != 0:
            for col, name in enumerate(extracted_data[0].keys()):
                # print(str(col + 1) + " " + name)
                curr_sheet.cell(column=col + 1, row=1, value=name)
                curr_sheet.cell(column=col + 1, row=1).font = Font(bold=True)
            row_num = 2
            for row in extracted_data:
                for col, value in enumerate(row.values()):
                    curr_sheet.cell(column=col + 1, row=row_num, value=value)
                row_num += 1


def workbook_documentation(in_file, output_dir, style_guide=None):
    """initialize the class and call output"""
    start_time = time.perf_counter()

    logging.info("Starting to process %s", in_file)

    documentation = WorkbookDocumentation(in_file, style_guide)

    # try:
    #     documentation = WorkbookDocumentation(in_file)
    # except Exception as e:
    #     error_message = "Unable to process {}\r\nException: {}".format(in_file, e)
    #     if __name__ != "__main__":
    #         # assumes being called by the tkinter program, so GUI error is OK.
    #         messagebox.showerror("Error", error_message)
    #     else:
    #         # print the error since it's run on the command line
    #         print(error_message)
    #     logging.error("Unable to process {} Exception: {}".format(in_file, e))
    #     return

    # todo - try/except? doesn't seem to be the area where errors occur.
    documentation.write_documentation(output_dir)

    logging.info("Finished processing %s", in_file)
    logging.info(
        "Processing took %s seconds", str(round(time.perf_counter() - start_time, 2))
    )


def main():
    """Process files without the GUI"""
    logging.basicConfig(
        filename="WorkbookDocumentation.log",
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    # try:
    # workbook_documentation("C:\\Users\\jj2362\\Desktop\\excel_test.twb",
    #                        'c:\\users\\jj2362\\desktop\\docs out')
    # workbook_documentation("C:\\Users\\jj2362\\Desktop\\HCC Opportunity Updated.twb",
    #                        'c:\\users\\jj2362\\desktop\\docs out')
    # workbook_documentation("C:\\Users\\jj2362\\Desktop\\1st Q1.twb",
    #                        'c:\\users\\jj2362\\desktop\\docs out')
    # workbook_documentation("C:\\Users\\jj2362\\Desktop\\docs in\\standard frequent flyer.tds",
    #                        'c:\\users\\jj2362\\desktop\\docs out')
    # workbook_documentation("/Users/jj2362/Desktop/CHLA performance overview.twbx",
    #                        '/Users/jj2362/Desktop/docs out',
    #                        './validator/HealtheAnalytics_style_guide.json')
    # workbook_documentation("/Users/jj2362/Desktop/Readmissions Discovery revised.twb",
    #                        '/Users/jj2362/Desktop/docs out',
    #                        './validator/HealtheAnalytics_style_guide.json')
    # workbook_documentation("./validator/tests/example_workbook.twb",
    #                        "/Users/jj2362/Desktop/docs out",
    #                        "./validator/tests/example_style_guide.json")
    # workbook_documentation("./validator/tests/example_workbook.twb",
    #                        "/Users/jj2362/Desktop/docs out",
    #                        "./validator/HealtheAnalytics_style_guide.json")
    workbook_documentation(
        "/Users/jj2362/Downloads/PROCUREMENT - Purchasing Dashboard.twbx",
        "/Users/jj2362/Desktop/docs out",
    )
    # except:
    #     logging.exception('')


if __name__ == "__main__":
    main()
